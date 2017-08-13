# coding:utf-8
import os
import time
import sqlite3
import difflib
from pyh import *
from robot.api import logger
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
class get_pm_task(object):
    # def __init__(self, DB_path ):
    #     self.DB_path = DB_path
    
    # @staticmethod
    # def get_pm_data(umac_dut, save_path):
    #     logger.info('start get current alarm data')
    #     t = time.time()
    #     cmd = 'SHOW PMSYSTASK;'
    #     cmd_result = umac_dut.execute_command(cmd)
    #     s1 = []
    #     result = cmd_result.return_string.split(' ')
    #     for item in result:
    #         print item
            # i = item.split(r'"-"')
            # s1.append(i[3]), s1.append(i[4])

        # get_pm_task.WriteInExcel(s1, save_path)

    @staticmethod
    def WriteInExcel(DB_path, save_path):
        conn = sqlite3.connect(DB_path)
        c = conn.cursor()
        alltable = c.execute('SELECT id, name, period, state from t_pm_task ').fetchall()
        logger.info('数据库文件读取完毕')
        wb = Workbook()
        ew = ExcelWriter(workbook = wb)
        ws = wb.worksheets[0]
        ws.title = "performance task"
        ws.cell(row=0,column=0).value = u'id'
        ws.cell(row=0,column=1).value = u'name'
        ws.cell(row=0,column=2).value = u'period'
        ws.cell(row=0,column=3).value = u'state'

        r = 1
        for item in alltable:
            ws.cell(row=r,column=0).value = item[0]
            ws.cell(row=r,column=1).value = item[1]
            ws.cell(row=r, column=2).value = item[2]
            ws.cell(row=r, column=3).value = item[3]
            r += 1

        ew.save(filename = save_path)
        c.close()
        conn.close()

    @staticmethod
    def ReadFromExcel(path):
        wb = load_workbook(path)
        ws = wb.get_sheet_by_name('performance task')
        s1 = []
        for i in range(1, len(ws.rows)):
            s2 = ''
            s2+=str(ws.cell(row=i,column=0).value)
            s2+=';'+ws.cell(row=i,column=1).value
            s2+=';'+str(ws.cell(row=i,column=2).value)
            s2+=';'+str(ws.cell(row=i,column=3).value)
            s1.append(s2)
        return s1

    #创建页面模板
    @staticmethod
    def html_template(before_path, after_path, s1, s2):
        save_path1 = before_path + '\\before.html'
        save_path2 = after_path + '\\after.html'

        if os.path.exists(save_path1):
            os.remove(save_path1)
        if os.path.exists(save_path2):
            os.remove(save_path2)
 
        page = PyH('summary result')
        page<<h2('3.performance task comparison:')
        div1 = page<<div(id='div1')
        table1 = div1<<table(border='2',id='pm_task')
        headtr = table1<<tr(id='headline', bgColor="#836FFF")<<th('compared result')<<th('before upgrade')<<th('after upgrade')
        
        
        page2 = PyH('result')
        page2<<h1('performance task', align='center')
        div2 = page2<<div(id='div1')
        table2 = div2<<table(border='2',id='pm_task')
        headtr = table2<<tr(id='headline')<<th('id')<<th('name')<<th('period')<<th('state')
        for i in s1:
            table2<<tr(td(i.split(';')[0]) + td(i.split(';')[1]) + td(i.split(';')[2]) + td(i.split(';')[3]))
        page2.printOut(save_path1)

        page3 = PyH('result')
        page3<<h1('performance task', align='center')
        div3 = page3<<div(id='div1')
        table3 = div3<<table(border='2',id='pm_task')
        headtr = table3<<tr(id='headline')<<th('id')<<th('name')<<th('period')<<th('state')
        for i in s2:
            table3<<tr(td(i.split(';')[0]) + td(i.split(';')[1]) + td(i.split(';')[2]) + td(i.split(';')[3]))
        page3.printOut(save_path2)
        return page, table1



    @staticmethod
    def pmtask_compare(save_path, before_path, after_path, save_path1, save_path2):
        try:
            s1 = get_pm_task.ReadFromExcel(save_path1)
        except IOError:
            logger.info('升级前没有生成性能任务文件')
        try:
            s2 = get_pm_task.ReadFromExcel(save_path2)
        except IOError:
            logger.info('升级后没有生成性能任务文件')


        page, tb = get_pm_task.html_template(before_path, after_path, s1, s2)
        change = 0
        for i in s1:
            if i not in s2:
                tb<<tr(td('deleted item') + td(i) + td('None'), bgColor='#F08080')
                change+=1

        for i in s2:
            if i not in s1:
                tb<<tr(td('new item')+ td('None') + td(i), bgColor='#00EE76')
                change+=1

        if change == 0:
            tb<<tr(td('no change') + td(a('before upgrade data file',target='_blank', href=before_path+'\\before.html')) + td(a('after upgrade data file',target='_blank', href=after_path+'\\after.html'))) 
        else:
            tb<<tr(td(' ') + td(a('before upgrade data file',target='_blank', href=before_path+'\\before.html')) + td(a('after upgrade data file',target='_blank', href=after_path+'\\after.html'))) 
        
        #!!!!!此处需注意 pyh.py源码中默认文件操作方式为'w',此处'a'
        #为修改源码后的参数，方便追加输出。  
        page.printOut(save_path, 'a')

        # wb = Workbook()
        # ew = ExcelWriter(workbook=wb)
        # ws = wb.worksheets[0]
        # ws.title = "compare result"
        # ws.cell(row=0, column=0).value = u'NEW'
        # ws.cell(row=0, column=1).value = u'DECREASED'

        # r = 1
        # for i in s1:
        #     for j in s2:
        #         if i not in s2:
        #             ws.cell(row=r,column=1).value = 'id: ' + str(i[0])+ ' name:' + str(i[1]) + ' period:' + str(i[2]) + ' state:' + str(i[3])
        #             r += 1
        #         if j not in s1:
        #             ws.cell(row=r,column=0).value = 'id: ' + str(j[0])+ ' name:' + str(j[1]) + ' period:' + str(j[2]) + ' state:' + str(j[3])
        #             r += 1

        # ew.save(filename=save_path)
        # d = difflib.HtmlDiff(wrapcolumn=80)
        # dd = d.make_file(s1, s2, context=True).replace("charset=ISO-8859-1","charset=UTF-8",1)
        # with open(save_path, 'a') as f:
        #     f.write('\n***性能任务对比***\n' + dd)
