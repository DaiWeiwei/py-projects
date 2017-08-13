# -*- coding: utf-8 -*- 

import umacHtmlQueryStruct
import umacHtmlCondition
from openpyxl.reader.excel import load_workbook
from openpyxl import workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.workbook  import  Workbook
import sqlite3
import re
import logging
import umacHtmlDB


class umacHtmlExcel:
    
    srcExcelFilePath= ''
    destExcelFilePath = ''
    dbFilePath = ''
    queryStructList = []
    allLeftTable = []
    wb = None
    excelLog = None

    def __init__(self,srcExcelFilePath,dbFilePath,destExcelFilePath):
        self.srcExcelFilePath = srcExcelFilePath
        self.dbFilePath = dbFilePath
        self.destExcelFilePath = destExcelFilePath
        self.queryStructList = []
        self.excelLog = logging.getLogger('umacHtmlMain.umacHtmlExcel')
        self.open_tables = []
        self.all_talbes = []

    @property
    def all_tables(self):
        return self.all_tables

    @all_tables.setter
    def all_tables(self,value):
        self.all_tables = value

    def readFromExcel(self,netType):
        wb = load_workbook(filename = self.srcExcelFilePath)
        #v = wb.get_sheet_names()
        ws = wb.get_sheet_by_name(netType)
        for rx in range(1,len(ws.rows)):
            queryStruct = umacHtmlQueryStruct.umacHtmlQueryStruct()
            queryStruct.netType = netType
            queryStruct.requiredNum =  ws.cell(row=rx,column=0).value
            queryStruct.version     =  ws.cell(row=rx,column=1).value
            queryStruct.functionName = ws.cell(row=rx,column=2).value
            queryStruct.functionDesc = ws.cell(row=rx,column=3).value
            queryStruct.opFlag = ws.cell(row=rx,column=4).value
            queryStruct.conditionNum = ws.cell(row=rx,column=5).value
            queryStruct.conditionList = []
            for col in range(queryStruct.conditionNum):
                conStruct = umacHtmlCondition.umacHtmlCondition()
                conStruct.tableName = ws.cell(row=rx,column=4*col+6).value
                conStruct.fieldName = ws.cell(row=rx,column=4*col+7).value
                conStruct.condition = ws.cell(row=rx,column=4*col+8).value
                conStruct.expResult = ws.cell(row=rx,column=4*col+9).value
                queryStruct.conditionList.append(conStruct)
            self.queryStructList.append(queryStruct)

    def readFromDB(self):
        conn = sqlite3.connect(self.dbFilePath)
        cu = conn.cursor()
        allTable = cu.execute("select name from sqlite_master where type = 'table' order by name").fetchall() #查询所有表
        self.allLeftTable = []
        for i in allTable:
            self.allLeftTable.append(i[0])
        
        querySql = ''
        for x in self.queryStructList:
            for y in x.conditionList:
                y.queResult = []
                if y.tableName is None:
                    print 'tableName is Null in excel file!'
                    x.finalResult = 'unknown'
                    continue
                try:
                    self.allLeftTable.remove(y.tableName)
                except Exception,ex:
                    pass
                    #self.excelLog.info(ex)
                    
                querySql = y.get_query_sql()
                try:
                    cu.execute(querySql)
                except Exception,ex:
                    regexp1 = 'no such table'
                    regexp2 = 'no such column'
                    p1 = re.compile(regexp1,re.I)
                    p2 = re.compile(regexp2,re.I)
                    if p1.search(str(ex)):
                        #self.excelLog.info(ex)
                        x.finalResult = 'Null'
                    elif p2.search(str(ex)):
                        #self.excelLog.info(ex)
                        #self.excelLog.info(querySql)
                        x.finalResult = 'unknown'
                    else:
                        #self.excelLog.info(ex)
                        #print 'error',y.tableName,y.fieldName
                        x.finalResult = 'unknown'
                else:  #数据库查询结果处理
                    r = cu.fetchall()
                    if len(r) == 0:
                        y.queResult.append('null')
                    if len(r) == 1:
                        if y.fieldName is None:
                            y.queResult.append('notnull')
                        elif r[0][0]=='':
                            y.queResult.append('null')
                        elif r[0][0]== y.fieldName:  #查询语句返回值为字段值，则认定为字段不存在
                            x.finalResult = 'unknown'
                            #self.excelLog.info('no such column:'+y.fieldName)
                        else:
                            value = r[0][0]
                            if '&' in str(value):  #返回结果中包含&，切片
                                valueList = str(value).split('&')
                                for vl in valueList:
                                    y.queResult.append(vl.lower().strip())       
                            else:
                                y.queResult.append(value.lower())
                    if len(r) >= 2:
                        if y.fieldName is None:
                            y.queResult.append('notnull')
                        else:
                            for i in range(len(r)):
                                y.queResult.append(str(r[i][0]).lower())
        cu.close()
        conn.close()

    def compare(self):
        self.open_tables = []

        for x in self.queryStructList:
            if x.finalResult is 'unknown':
                x.finalResult = 'Closed'  #未知情况的被判定为未开启
                continue
            
            if x.opFlag == 0:
                x.finalResult = 'Open'
            else:
                x.finalResult = 'Closed'
            
            for y in x.conditionList:
                if x.opFlag == 0:
                    if '!' in str(y.expResult):
                        exp = str(y.expResult).replace('!','')
                        if exp in y.queResult:
                            x.finalResult = 'Closed'
                    else:
                        if str(y.expResult).lower() not in  y.queResult:                 
                            x.finalResult = 'Closed'
                
                elif x.opFlag == 1:
                    if '!' in str(y.expResult):
                        exp = str(y.expResult).replace('!','')
                        if exp not in y.queResult:
                            x.finalResult = 'Open'
                    else:
                        if str(y.expResult).lower() not in  y.queResult:                 
                            x.finalResult = 'Open'
            if x.finalResult == 'Open':
                self.open_tables.extend([c.tableName.replace('_',' ') for c in x.conditionList])
    def max(self,a,b):
        if a > b:
            return a
        else:
            return b

    def writeToExcel(self,netType):
        ew = ExcelWriter(workbook = self.wb)
        ws=self.wb.create_sheet()
        ws.title =  netType
        ws.cell(row=0,column=0).value =  u'需求编号'
        ws.cell(row=0,column=1).value =  u'版本号'
        ws.cell(row=0,column=2).value =  u'功能名称'
        ws.cell(row=0,column=3).value =  u'功能描述'
        ws.cell(row=0,column=4).value =  u'当前状态'
        
        #不需要输出具体开关
        #maxConNum = 0
        #for x in self.queryStructList:
        #    maxConNum = self.max(maxConNum,x.conditionNum)
        #for i in range(maxConNum):
        #    ws.cell(row=0,column=2+i).value =  'Sw'+str(i+1)

        r = 1
        for x in self.queryStructList:
            if x.netType is netType:
                ws.cell(row=r,column=0).value = x.requiredNum
                ws.cell(row=r,column=1).value = x.version
                ws.cell(row=r,column=2).value = x.functionName
                ws.cell(row=r,column=3).value = x.functionDesc
                ws.cell(row=r,column=4).value = x.finalResult
            
            #不需要输出具体开关
            #col = 2
            #for y in x.conditionList:
            #    ws.cell(row=r,column=col).value = y.queResult
            #    col = col+1
                r = r+1
        ew.save(filename = self.destExcelFilePath)


    def writeALlLeftTable(self,allLeftTable):
        umacDB = umacHtmlDB.umacHtmlDB()
        conn = umacDB.get_conn(self.dbFilePath)
        ew = ExcelWriter(workbook = self.wb)
        ws=self.wb.create_sheet()
        ws.title =  u'未分析命令'
        r = 0
        for x in allLeftTable:
            allStr = ''
            sql = 'select * from '+x
            tc = umacDB.fetchall(conn, sql) #获取表数据
            tf = umacDB.get_table_field(conn,x)#获取表头
            
            for t in tf:
                allStr = allStr + str(t) + '    '
            allStr = allStr + '\r\n'
            
            #tcList = self.column_cover_to_line(tc,length)
            
            #tfList = self.tupleList_cover_to_list(tf)
            
            tcStr = self.formate(tc)
            allStr = allStr +tcStr
            x = x.replace('_', ' ')
            ws.cell(row=r,column=0).value = x
            ws.cell(row=r,column=1).value = allStr
            r = r + 1
#             for i in range(len(tfList)):
#                 ws.cell(row=r,column=i+1).value = tfList[i]
#             r = r+1
#             
#             for j in range(len(tcList)):
#                 ws.cell(row=r,column=j+2).value = tcList[j]
#             r = r+1 
            
        ew.save(filename = self.destExcelFilePath)
        
    def column_cover_to_line(self,tc,length): 
        tcList = []
        for l in range(length):
            tcStr = ''
            for t in tc:
                tcStr = tcStr+str(t[l])+'\r\n'
            tcList.append(tcStr)
        return tcList
    
    def list_cover_to_str(self,tc):
        pass
    
    def tupleList_cover_to_list(self,tf):
        tflist = []
        for t in tf:
            tflist.append(t[0])
        return tflist
    
    def formate(self,tc):
        allStr = ''
        for x in tc:
            for y in x:
                allStr = allStr + str(y) + '    '
            allStr = allStr + '\r\n'
        return allStr
    def execute(self):
        self.wb = Workbook()
        wbx = load_workbook(filename = self.srcExcelFilePath)
        netTypeList = wbx.get_sheet_names()
        netTypeList.remove(u'命令过滤')
        netTypeList.remove(u'规则表命令')
        for netType in netTypeList:
            self.readFromExcel(netType)
        self.readFromDB()
        self.compare()

        self.writeALlLeftTable(self.allLeftTable)
        for netType in netTypeList:   
            self.writeToExcel(netType)

        self.open_tables = list(set(self.open_tables))
        self.writeAllTalbes()

    def writeAllTalbes(self):
        ew = ExcelWriter(workbook = self.wb)
        ws=self.wb.create_sheet()
        ws.title =  u'所有命令'
        all_tables = []
        open_tables = []
        close_tables = []
        for idx, x in enumerate(self.all_tables):
            if x in self.open_tables:
                open_tables.append([x,'Open'])
            else:
                close_tables.append([x,'Close'])
        all_tables = open_tables
        all_tables.extend(close_tables)
        for idx,table in enumerate(all_tables):
            ws.cell(row=idx,column=0).value = table[0]
            ws.cell(row=idx,column=1).value = table[1]
        ew.save(filename = self.destExcelFilePath)
if __name__ == '__main__':
    print ''
