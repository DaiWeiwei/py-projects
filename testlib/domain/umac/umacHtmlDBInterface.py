# -*- coding: utf-8 -*- 
import os
import logging
import umacHtml
import time
from openpyxl.reader.excel import load_workbook
from openpyxl import workbook
from openpyxl.writer.excel import ExcelWriter
import umacHtmlDB

class umacHtmlDBInterface:
    htmlFilePath = ''
    excelFilePath = ''
    dbFilePath = ''
    dbLog = None
    htmlData = ''

    def __init__(self,htmlFilePath,excelFilePath,dbFilePath):
        self.htmlFilePath = htmlFilePath
        self.excelFilePath = excelFilePath
        self.dbFilePath = dbFilePath
        self.dbLog = logging.getLogger('umacHtmlMain.umacHtmlDBInterface')
        self.htmlData = ''
        self.all_tables = []

    def create_all_table(self): 
        umacH = umacHtml.umacHtml(self.htmlData)
        allData = umacH.getAllCmdNameAndContent()
        umacDB = umacHtmlDB.umacHtmlDB()
        conn = umacDB.get_conn(self.dbFilePath)
        filterCommand = self.get_filter_command()
        dbTime = 0
        
        #从全表中删除过滤命令
        #20150528特殊表处理  Modify by zsc
        for com in filterCommand:
            temp = []
            for j in allData:
                if j[1].strip()!=com:
                    temp.append(j)
            allData = temp
        self.all_tables = allData
        for data in allData:
            table = data[1]
            table = table.strip()
            tableName = table.replace(' ','_')
            htmlTable = umacH.umacHtmlParase(data[3])
            if htmlTable is None:
                #.dbLog.info('No Found the table in HtmlFile:'+data[1])
                continue
            t0 = time.time()
            umacDB.create_table_fromHead(conn,tableName,htmlTable[0])
            for row in range(1,len(htmlTable)):
                umacDB.insert_table(conn,tableName,row,htmlTable[row])
            t1 = time.time()
            #self.dbLog.info(data[1]+' time='+str(t1-t0))
            dbTime = dbTime + t1 - t0
        conn.close()
        print 'dbTime=', dbTime
        #self.dbLog.info(dbTime)
        
    def create_partial_table(self):
        if os.path.exists(self.excelFilePath) and os.path.isfile(self.excelFilePath):
            dict1 = self.get_command_table()
        else:
            print('No such file or directory:[{}]'.format(self.excelFilePath))
            return
        
        dbTime = 0
        umacDB = umacHtmlDB.umacHtmlDB()
        conn = umacDB.get_conn(self.dbFilePath)
        umacH = umacHtml.umacHtml(self.htmlData)
        
        #20150528特殊表处理 Modify by zsc
#         neType = umacH.get_showCFGCmdName()
#         neType = neType.replace(' ','_')
        
        self.all_tables = dict1.keys()
        for key in dict1.keys():
            strTable = umacH.get_cmd_content(str(key)) #字典返回的对象是unicode
            if strTable is None:
                #self.dbLog.info('No Found the Command in HtmlFile:'+key)
                continue
            htmlTable = umacH.umacHtmlParase(strTable)
            if htmlTable is None:
                #self.dbLog.info('No Found the table in HtmlFile:'+key)
                continue
            t0 = time.time()
            umacDB.create_table_fromHead(conn,dict1[key],htmlTable[0])
            for row in range(1,len(htmlTable)):
                umacDB.insert_table(conn,dict1[key],row,htmlTable[row])
            t1 = time.time()
            dbTime = dbTime +t1-t0
        conn.close()
        print 'dbTime=',dbTime

    def get_command_table(self):
        dict1 = {}
        #print self.excelFilePath
        wb = load_workbook(filename = self.excelFilePath)
        ws = wb.get_sheet_by_name(u'规则表命令')
        for rx in range(1,len(ws.rows)):
            command = ws.cell(row=rx,column=1).value
            table = command.replace(' ','_')
            #print command,table
            dict1.setdefault(command,table)
        return dict1
    
    def get_filter_command(self):
        filterCommand = []
        wb = load_workbook(filename = self.excelFilePath)
        ws = wb.get_sheet_by_name(u'命令过滤')
        for rx in range(1,len(ws.rows)):
            command = ws.cell(row=rx,column=1).value
            filterCommand.append(command)
        return filterCommand
    
    def modify_special_table(self,ne):
        st = ne.replace(' ','_')
        wb = load_workbook(filename = self.excelFilePath)
        ew = ExcelWriter(workbook = wb)
        netTypeList = wb.get_sheet_names()
        netTypeList.remove(u'命令过滤')
        for netType in netTypeList:
            ws = wb.get_sheet_by_name(netType)
            for rx in range(0,len(ws.rows)):
                tableName = ws.cell(row=rx,column=5).value
                if tableName == 'SHOW_COMBOCFG':
                    ws.cell(row=rx,column=5).value = st
            ew.save(self.excelFilePath)
            
    def create_DB(self,tableType):
        if os.path.exists(self.htmlFilePath) and os.path.isfile(self.htmlFilePath):
            f = open(self.htmlFilePath)
            self.htmlData = f.read()
            f.close()
        else:
            print('No such file or directory:[{}]'.format(self.htmlFilePath))
            return

        if os.path.exists(self.dbFilePath):
            os.remove(self.dbFilePath)
        else:
            #print('No such file or directory:[{}]'.format(self.dbFilePath))
            print ''
        
        #20150528特殊表处理  Modify by zsc
        umacH = umacHtml.umacHtml(self.htmlData)
        ne = umacH.get_showCFGCmdName()
        if ne == 'SHOW SGSNCFG' or ne == 'SHOW MMECFG':
            self.modify_special_table(ne)

        if tableType is 'allTable':
            self.create_all_table()
            #print 'allTable'
        
        if tableType is 'partialTable':
            self.create_partial_table()
            
