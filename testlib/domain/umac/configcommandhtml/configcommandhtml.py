# coding:utf-8
import os
from openpyxl.reader.excel import load_workbook
from testlib.infrastructure.device.umac.umac import uMac
from testlib.infrastructure.iecapture.iecapture import IeCapture
from robot.api import logger

class ConfigCommandHtml(object):
    def __init__(self, umac_dut, rule_file):
        self.umac_dut = umac_dut
        self.rule_file = rule_file

    def _get_ignore_commands(self):
        wb = load_workbook(filename=self.rule_file)
        # v = wb.get_sheet_names()
        ws = wb.get_sheet_by_name(u'忽略命令规则')
        ignore_commands = []
        for rx in range(1, len(ws.rows)):
            command = ws.cell(row=rx, column=0).value
            if not command:
                continue
            ignore_commands.append(command)
        return ignore_commands

#获取特殊命令，这些命令在show cmd中不体现的，但是需要检查的，add by xjr 0606
    def _get_special_commands(self):
        wb = load_workbook(filename=self.rule_file)
        # v = wb.get_sheet_names()
        ws = wb.get_sheet_by_name(u'特殊命令')
        spec_commands = []
        for rx in range(1, len(ws.rows)):
            command = ws.cell(row=rx, column=0).value
            if not command:
                continue
            spec_commands.append(command)
        return spec_commands
#结束获取特殊命令！

    def _get_all_config_commands(self):
        logger.info('start get all config commands')
        if not self.umac_dut:
            raise Exception('umac dut is none,please check umac name first!')
        self.umac_dut.logout()
        self.umac_dut.login_with_try(20)
        logger.info('start get all config commands:send SHOW CMD:LEVEL="SM_SYSTEMLEVEL_4";')
        cmd_result = self.umac_dut.execute_command('SHOW CMD:LEVEL="SM_SYSTEMLEVEL_4";', 300)
        # print cmd_result.return_string
##        with open('c:/umac_config.txt','wb') as f:
##          f.write(cmd_result.return_string)
        show_commands = []
        #特殊处理一下，命令解释行中有&的先替换掉 20160901  xjr
        cmd_result.return_string=cmd_result.return_string.replace('ICMPV6&ND','ICMPV6ND')
        for line in cmd_result.return_string.split('&'):
##            print line.split('"-"')[1].strip('"')
            show_commands.append(line.split('"-"')[1].strip('"'))
        self.umac_dut.logout()
        return show_commands

    def _get_config_commands_to_be_run(self):
        all_config_commands = self._get_all_config_commands()
        ignore_commands = self._get_ignore_commands()
        for cmd in ignore_commands:
            if cmd in all_config_commands:
                all_config_commands.remove(cmd)
     #增加特殊命令到运行命令中 addby xjr 0606
        spec_commands=self._get_special_commands()
        for cmds in spec_commands:
            all_config_commands.append(cmds)
        return all_config_commands

 #增加内部软参命令参数spec_cmd，用于查询内部软参add by xjr 20160615
    def get_config_commands_and_save(self,current_local_save_path,spec_cmd):
        all_config_commands = self._get_config_commands_to_be_run()
        all_config_commands.append(spec_cmd) #添加查询内部软参的命令add by xjr 20160615
        file_name = "{0}/{1}".format(current_local_save_path,'umac_config_commands.txt')
        with open(file_name, 'wb') as f:
            f.write('\n'.join(all_config_commands))
        return file_name

    def get_config_commands_html(self, command_file):
        if not os.path.exists(command_file):
            raise Exception('command file not exists:{0}'.format(command_file))
        IeCapture.generate_umac_config_command_html(command_file)


if __name__ == '__main__':
    umac = uMac('192.0.19.10', '7722', 'admin', '', 19)
    rule_file = 'E:/umac_local/umac_html_rule.xlsx'
    cch = ConfigCommandHtml(umac, rule_file)
    print '*' * 8
    #print cch._get_config_commands_to_be_run()
    with open('c:/ttte.txt','wb') as f:
        f.write('\n'.join(cch._get_config_commands_to_be_run()))
    # print cch._get_ignore_commands()
