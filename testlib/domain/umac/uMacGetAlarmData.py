#coding:utf-8
import os
import time
import difflib
from pyh import *
from robot.api import logger
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
import sys

reload(sys)
sys.setdefaultencoding('utf8')

class GetAlarmData(object):
    # def __init__(self, umac_name, upgrade):
    #     self.before_updrade_s1 = []
    #     self.after_updrade_s2 = []
    #     self.before_updrade_s1 = []
    #     self.after_updrade_s2 = []
    #     self.umac_name = umac_name
    #     self.upgrade = upgrade

    @staticmethod
    def get_alarm_data(umac_dut, save_path):
        logger.info('start get current alarm data')
        # upgrade = int(upgrade)
        t = time.time()

        cmd = 'SHOW FAULTALARM;'
        cmd_result1 = umac_dut.execute_command(cmd)

        cmd = 'SHOW ALARMCODE;'
        cmd_result2 = umac_dut.execute_command(cmd)


        #告警码、位置写入列表
        s1 = []
        result1 = cmd_result1.return_string.split('&')
        for item in result1:
            i = item.split(r'"-"')
            s1.append(i[3]), s1.append(i[8])


        #告警码、描述写入字典
        dict1 = {}
        result2 = cmd_result2.return_string.split('&')
        for item in result2:
            i = item.split(r'"-"')
            dict1[i[1]] = i[5]

        #umac_config.delete_allfiles(save_path)

        GetAlarmData.WriteInExcel(s1, dict1, save_path)
        logger.info('end get current alarm data!, {0}s'.format(time.time() - t))

    @staticmethod
    def ReadFromExcel(path):
        wb = load_workbook(path)
        ws = wb.get_sheet_by_name('alarm data')
        s2 = []
        for i in range(1, len(ws.rows)):
            s1 = ''
            s1+=str(ws.cell(row=i,column=0).value)
            s1+=';'+ws.cell(row=i,column=1).value
            s1+=';'+ws.cell(row=i,column=2).value
            s2.append(s1)
        return s2

    @staticmethod
    def WriteInExcel(s1, dict1, save_path):
        
        wb = Workbook()
        ew = ExcelWriter(workbook = wb)
        ws = wb.worksheets[0]
        ws.title = "alarm data"
        ws.cell(row=0,column=0).value = u'alarm code'
        ws.cell(row=0,column=1).value = u'location'
        ws.cell(row=0,column=2).value = u'alarm description'
        # s2 = []
        r = 1
        for item1 in range(0, len(s1), 2):
            ws.cell(row=r,column=0).value = s1[item1]
            ws.cell(row=r,column=1).value = s1[item1+1]
            for item2 in dict1.keys():
                if item2 == s1[item1]:
                    ws.cell(row=r,column=2).value = dict1.get(item2)
                    # s2.append(dict1.get(item2))
            r += 1
        ew.save(filename = save_path)

    #创建页面模板
    @staticmethod
    def html_template(before_path, after_path, s1, s2):
        save_path1 = before_path + '\\before.html'
        save_path2 = after_path + '\\after.html'

        if os.path.exists(save_path1):
            os.remove(save_path1)
        if os.path.exists(save_path2):
            os.remove(save_path2)

        href_path = save_path1
        for i in range(3):
            href_path = os.path.split(href_path)[0]
        href_path = href_path + '\\output\\sig_comparison.html'

        page = PyH('summary result')
        page<<h1('Upgrade Report', align='center')
        page<<h2('1.sigtrace comparison:')
        page << div(b(a('>>>go to comparison result',target='_blank', href=href_path)))
        page<<h2('2.alarm comparison:')
        div1 = page<<div(id='div1')
        table1 = div1<<table(border='2',id='alarm')
        headtr = table1<<tr(id='headline', bgColor="#836FFF")<<th('compared result')<<th('before upgrade')<<th('after upgrade')


        page2 = PyH('result')
        page2<<h1('alarm data', align='center')
        div2 = page2<<div(id='div1')
        table2 = div2<<table(border='2',id='alarm')
        headtr = table2<<tr(id='headline')<<th('alarm code')<<th('location')<<th('description')
        for i in s1:
            table2<<tr(td(i.split(';')[0]) + td(i.split(';')[1]) + td(i.split(';')[2]))
        page2.printOut(save_path1)

        page3 = PyH('result')
        page3<<h1('alarm data', align='center')
        div3 = page3<<div(id='div1')
        table3 = div3<<table(border='2',id='alarm')
        headtr = table3<<tr(id='headline')<<th('alarm code')<<th('location')<<th('description')
        for i in s2:
            table3<<tr(td(i.split(';')[0]) + td(i.split(';')[1]) + td(i.split(';')[2]))
        page3.printOut(save_path2)

        return page, table1


    @staticmethod
    def alarm_compare(save_path, before_path, after_path, save_path1, save_path2):
        try:
            s1 = GetAlarmData.ReadFromExcel(save_path1)
        except:
            logger.info('升级前没有生成告警文件')
        try:
            s2 = GetAlarmData.ReadFromExcel(save_path2)
        except:
            logger.info('升级后没有生成告警文件')

        page, tb = GetAlarmData.html_template(before_path, after_path, s1, s2)
        change = 0
        for i in s1:
            if i not in s2:
                tb<<tr(td('deleted item') + td(i) + td('None'), bgColor='#F08080')
                change += 1

        for i in s2:
            if i not in s1:
                tb<<tr(td('new item')+ td('None') + td(i), bgColor='#00EE76')
                change += 1
        if change == 0:
            tb<<tr(td('no change') + td(a('before upgrade data file',target='_blank', href=before_path+'\\before.html')) + td(a('after upgrade data file',target='_blank', href=after_path+'\\after.html')))

        else:
            tb<<tr(td(' ') + td(a('before upgrade data file',target='_blank', href=before_path+'\\before.html')) + td(a('after upgrade data file',target='_blank', href=after_path+'\\after.html')))
        page.printOut(save_path)


        # wb = Workbook()

        # ew = ExcelWriter(workbook = wb)
        # ws = wb.worksheets[0]
        # ws.title = "compare result"
        # ws.cell(row=0,column=0).value = u'location'
        # ws.cell(row=0,column=1).value = u'new alarm'
        # ws.cell(row=0,column=2).value = u'decreased alarm'
        # ws.cell(row=0,column=3).value = u'different items'

        # r = 1
        # for i in range(0, len(s1), 2):
        #     if s1[i+1] not in s2:
        #         ws.cell(row=r,column=0).value = s1[i+1]
        #         ws.cell(row=r,column=2).value = 'alarm code: ' + s1[i] + ';  ' + 'description: ' + s3[i/2]
        #         r+=1
        #         continue

        #     if s2[i+1] not in s1:
        #         ws.cell(row=r,column=0).value = s2[i+1]
        #         ws.cell(row=r,column=1).value = 'alarm code: ' + s2[i] + ';  ' + 'description: ' + s4[i/2]
        #         r+=1
        #         continue

        #     for j in range(0, len(s2), 2):
        #         #相同位置判断告警变化
        #         if s1[i+1] == s2[j+1] and not s1[i] == s2[j]:
        #             if s1[i] not in s2:
        #                 ws.cell(row=r,column=0).value = s1[i+1]
        #                 ws.cell(row=r,column=2).value = 'alarm code: ' + s1[i] + ';  ' + 'description: ' + s3[i/2]
        #                 r+=1
        #                 continue

        #             elif s2[j] not in s1:
        #                 ws.cell(row=r,column=0).value = s2[j+1]
        #                 ws.cell(row=r,column=1).value = 'alarm code: ' + s2[j] + ';  ' + 'description: ' + s4[j/2]
        #                 r+=1
        #                 continue
        #                 #判断是否因为数据存入顺序不同导致的比对差异
        #             else:
        #                 for k in range(0, len(s2), 2):
        #                     if s1[i] == s2[k] and s1[i+1] == s2[k+1]:
        #                         break
        #                     #最后一次无论是否成功都跳出循环，需要另外判断
        #                     if k == len(s2) - 2:
        #                         if not s1[i] == s2[k] or not s1[i+1] == s2[k+1]:
        #                             ws.cell(row=r,column=0).value = s2[j+1]
        #                             ws.cell(row=r,column=3).value = 'before:{ alarm code : ' + s1[i] + 'description : ' + s3[i/2] + ' } '+ ' after:{ alarm code : ' + s2[j] + 'description : ' + s4[j/2] + ' } '
        #                             r+=1


        # ew.save(filename = save_path)
        # d = difflib.HtmlDiff(wrapcolumn=60)
        # dd = d.make_file(s1, s2, context=True).replace("charset=ISO-8859-1","charset=UTF-8",1)
        # with open(save_path, 'w') as f:
        #     f.write('***告警对比***\n' + dd)














