#coding:utf-8
import os
import re
import xlrd
from openpyxl import load_workbook

from filterresultanalysis import FilterResultAnalysis
from umacfunctioncharactor import uMacFunctionCharactor
from umacconfigcmdmanagement import UmacConfigCmdManagement
from uMacLicense import uMacLicense
from testlib.infrastructure.utility.mylog import create_log,get_log,release_log

COL_REQUEST_NO   = 0
COL_VERSION      = 1
COL_FUNCTIONNAME = 2
COL_DESC         = 3
COL_CMD          = 4

class uMacFunctionCharactorRefresh:
    def __init__(self,all_config_cmd_list,
                      filter_matched_cmd_list,
                      function_charactor_file,
                      ne_type,
                      output_dir):
        self.log_out = get_log()
        self._output_dir = output_dir

        self._dict_ne_type = {'MME':['MME',u'公共'],'SGSN':['SGSN',u'公共'],'COMBO':['MME','SGSN',u'公共']}
        self._sheet_names = ['MME','SGSN',u'公共']

        self.ne_sheet_names = self._dict_ne_type.get(ne_type.upper(),[])
        if not self.ne_sheet_names:
            raise Exception('cannot find sheet name {0} in file {1}'.format(ne_type,function_charactor_file))

        self._umac_config_cmd_management = UmacConfigCmdManagement()
        self._umac_all_config_cmd_management = UmacConfigCmdManagement()

        self._function_charactor_file = function_charactor_file

        for cmd in filter_matched_cmd_list:
            if not self._umac_config_cmd_management.append_cmd(cmd):
                self.log_out.error('append cmd to filter matched cmd list fail:{0}'.format(cmd))

        for cmd in all_config_cmd_list:
            if not self._umac_all_config_cmd_management.append_cmd(cmd):
                self.log_out.error('append cmd to all config cmd list fail:{0}'.format(cmd))

        self.ne_funcition_charactor = self.get_ne_funcition_charactor()
        self.analysis_function_charactor()
        self.print_function_charactor_result()
        self.update_function_charactor_to_file()

    def get_ne_funcition_charactor(self):
        ne_funcition_charactor = {}
        work_book = xlrd.open_workbook(self._function_charactor_file)#,formatting_info = True
        for sheet_name in self.ne_sheet_names:
            ne_funcition_charactor[sheet_name] = self.read_sheet(work_book,sheet_name)
        return ne_funcition_charactor

    def _print_cells(self, sheet_name, cells):
        ls = []
        self.log_out.debug('',flag = 1)
        ls.append(u'  cell pos of sheet {0}'.format(sheet_name))
        for cell in cells:
            ls.append('{0}{1}'.format(' '*4,cell))
        self.log_out.debug('\n'.join(ls),flag = 0)

    def read_sheet(self, work_book, sheet_name):
        sheet = work_book.sheet_by_name(sheet_name)
        if not sheet:
            return
        sheet_function_charactor = []
        merged_cells = sheet.merged_cells
        #print 'merged_cells:',merged_cells
        cells = self._get_cells_pos_of_function_name(sheet)
        self._print_cells(sheet_name, cells)

        for cell in cells:
            cmd_list = []
            #如果是合并单元格，则其值在单元格的第一行
            function_name = sheet.row(cell[0])[COL_FUNCTIONNAME].value
            if not function_name:
                continue
            for x in range(cell[0], cell[1]):
                cell_value = sheet.row(x)[COL_CMD].value.lower().replace(';'," ").replace('"',"").replace("'","")
                cmd_list.append(cell_value)
            function_charactor = uMacFunctionCharactor(function_name,cell,cmd_list)
            sheet_function_charactor.append(function_charactor)
        return sheet_function_charactor

    def update_function_charactor_to_file(self):
        work_book  = load_workbook(self._function_charactor_file)
        for sheet_name in self.ne_sheet_names:
            sheet = work_book.get_sheet_by_name(sheet_name)
            if not sheet:
                continue
            sheet_function_charactor = self.ne_funcition_charactor.get(sheet_name,None)
            if not sheet_function_charactor:
                self.log_out.error('cannot find function charactor by sheet name {0}'.format(sheet_name))
                continue
            self.update_sheet_function_charactor(sheet,sheet_function_charactor)
        work_book.save(u'{0}/umac功能特性表.xlsx'.format(self._output_dir))

    def update_sheet_function_charactor(self,sheet,sheet_function_charactor):
        for function_charactor in sheet_function_charactor:
            sheet.cell(row=function_charactor.cell_pos[0],column=COL_DESC).value = function_charactor.function_open_status
            if function_charactor.cell_pos[0] + 1 >= function_charactor.cell_pos[1]:
                continue
            for col in range(1,COL_DESC+1):
                sheet.merge_cells(None,
                                  function_charactor.cell_pos[0],
                                  col,
                                  function_charactor.cell_pos[1]-1,
                                  col)

    def analysis_function_charactor(self):
        for sheet_name, sheet_function_charactor in self.ne_funcition_charactor.items():
            for function_charactor in sheet_function_charactor:
                function_charactor.function_open_status = self._analysis_function_charactor_open_status(function_charactor)

    def print_function_charactor_result(self):
        self.log_out.debug('',flag = 1)
        self.log_out.debug('  print function charactor result:',flag = 0)
        space_str = ' '*4
        for sheet_name, sheet_function_charactor in self.ne_funcition_charactor.items():
            self.log_out.debug(u'{0}sheet_name:{1}'.format(space_str,sheet_name), flag = 0)
            for function_charactor in sheet_function_charactor:
                if not function_charactor.function_open_status:
                    continue
                self.log_out.debug(u'{0}'.format(function_charactor),flag=0)
                self.log_out.debug(u'{0}{1}'.format(space_str,'-'*10),flag = 0)

    def _get_sheet_index_by_name(self,sheet_name):
        return self._sheet_names.index(sheet_name)

    def _analysis_function_charactor_open_status(self,function_charactor):
        self.log_out.debug('-->{0} _is_in_all_config_cmd_list '.format(function_charactor._name),0)
        if not self._is_in_all_config_cmd_list(function_charactor):
            return  u'功能没开启'
        self.log_out.debug('-->{0} _is_in_matched_case_cmd_list '.format(function_charactor._name),0)
        if self._is_in_matched_case_cmd_list(function_charactor):
            return u'功能开启有用例'
        return  u'功能开启没用例'

    def _is_in_all_config_cmd_list(self,function_charactor):
        return self._is_function_charactor_cmd_in(self._umac_all_config_cmd_management,function_charactor)

    def _is_in_matched_case_cmd_list(self,function_charactor):
        return self._is_function_charactor_cmd_in(self._umac_config_cmd_management,function_charactor)

    def _is_function_charactor_cmd_in(self,umac_config_cmd_management,function_charactor):
        result = -1
        output_flag = 0
        if function_charactor._name == u'签约通配，APN解析失败后用本地默认APN再次解析':
            output_flag = 1
        for cmd in function_charactor.cmd_list:
            if not cmd:
                continue
            result = 0
            for sub_cmd in cmd:
                result += umac_config_cmd_management.is_cmd_exists(sub_cmd)
                if output_flag:
                    self.log_out.debug('-->{0}:{1}'.format(sub_cmd,result),0)
                if result:
                    break
            if not result:
                return False
        if result == -1:
            return False
        return True
    def _get_cells_pos_of_function_name(self,sheet):
        return self._get_col_cells_pos(sheet,COL_FUNCTIONNAME)

    @staticmethod
    def _sort_cells_pos(cells_pos_of_function_name):
        cells_pos_of_function_name.sort(lambda x1,x2:cmp(x1[0],x2[0]))

    def _get_merged_cells_flag(self,merged_cells):
        merged_cells_flag = []
        if len(merged_cells) > 0:
            merged_cells_flag = [0] * (merged_cells[-1][1])
        x = len(merged_cells_flag)
        for merge_cell in merged_cells:
            for row in range(merge_cell[0],merge_cell[1]):
                merged_cells_flag[row] = 1
        return merged_cells_flag

    def _get_col_cells_pos(self,sheet,col):
        merged_cells = self._get_merged_cells_of_row(sheet,col,col + 1)
        self._sort_cells_pos(merged_cells)

        merged_cells_flag = self._get_merged_cells_flag(merged_cells)

        row_num = sheet.nrows
        col_cells = []
        for row in range(1,row_num):
            if self._is_merged_cell(merged_cells_flag,row):
                continue
            col_cells.append( (row,row+1,col,col+1))
        col_cells.extend(merged_cells)
        self._sort_cells_pos(col_cells)
        return col_cells

    @staticmethod
    def _is_merged_cell(merged_cells,row):
        try:
            return merged_cells[row]
        except:
            return False

    @staticmethod
    def _get_merged_cells_of_row(sheet,col,col_range):
        #行合并
        row_merged_cells = []
        #merged_cell = (row,row_range,col,col_range) 从row 到row_range-1，不包含row_range
        for merged_cell in sheet.merged_cells:
            #get_log().info(">>"+str(merged_cell),flag = 0)
            if merged_cell[2] == col and merged_cell[3] == col_range:
                row_merged_cells.append(merged_cell)
        return row_merged_cells


def parse_batch_cmd_file(batch_cmd_file):
    with open(batch_cmd_file,'r') as f:
        lines = f.read().lower().replace(';'," ").replace('"',"").replace("'","").splitlines()
##        result = re.findall('^[^-].+',content,re.M)
##        if not result:
##            return []
        result = [ r.strip(" \r\n") for r in lines if len(r) > 0 and not r.startswith('--')]
        return result

def refresh_function_charactor(umac_config):
    batch_cmd_file = '{0}/Batchcommandmml.txt'.format(umac_config.run_save_path)
    if not os.path.exists(batch_cmd_file):
        return

    create_log(umac_config.log_path)
    #try:
    filter_result_file ='{0}/umac_filter_log.txt'.format(umac_config.temp_output_path)

    filter_matched_cmds = FilterResultAnalysis(filter_result_file).get_result()

    batch_cmd_list = parse_batch_cmd_file(batch_cmd_file)

    analysis_license_file = '{0}/analysis_license.txt'.format(umac_config.cmd_license_dir(0))

    license_cmd_list = uMacLicense().parse_license_file(analysis_license_file)

    function_charactor_rule_file = u'{0}/合并中文规则表-NEW.xlsx'.format( umac_config.local_save_path)
    all_config_cmd_list = []
    all_config_cmd_list.extend(batch_cmd_list)
    all_config_cmd_list.extend(license_cmd_list)
    uMacFunctionCharactorRefresh(all_config_cmd_list,
                                 filter_matched_cmds,
                                 function_charactor_rule_file,
                                 umac_config.ne_type,
                                 umac_config.output_dir)
##    except Exception,e:
##        print 'refresh_function_charactor fail:{0}'.format(e)
##        raise e

    release_log()


if __name__ == '__main__':
    create_log(r'D:\project_version_valid')
    filter_result_file =r'D:\project_version_valid\umac\combo_mmegngp_sgsn_27\temp_output\umac_filter_log.txt'
    filter_matched_cmds = FilterResultAnalysis(filter_result_file).get_result()

    batch_cmd_file =   r'D:\project_version_valid\umac\combo_mmegngp_sgsn_27\Batchcommandmml.txt'
    batch_cmd_list = parse_batch_cmd_file(batch_cmd_file)

    analysis_license_file = r'D:\project_version_valid\umac\combo_mmegngp_sgsn_27\before_upgrade\license\analysis_license.txt'
    license_cmd_list = uMacLicense().parse_license_file(analysis_license_file)

    all_config_cmd_list = []
    all_config_cmd_list.extend(batch_cmd_list)
    all_config_cmd_list.extend(license_cmd_list)
    uMacFunctionCharactorRefresh(all_config_cmd_list,
                                 filter_matched_cmds,
                                 u'D:\\project_version_valid\\umac\\合并中文规则表-NEW.xlsx',
                                 'mme',
                                 "c:")

    release_log()
