# -*- coding: utf-8 -*-

## set the file encoding as utf-8
import sys
reload(sys)
sys.setdefaultencoding('utf-8')

import xlrd
import xlwt
import re
import os
import operator

#使用openpyxl写xlsx，xlwt不支持xlsx
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import Cell

class xlsOP:
    '''
    改造的基本思路：
    （1）__init__中传入多个需要处理的sheetnames，比如SGSN、MME
    （2）parseRuleXls修改为parseSingleSheet，处理单个sheet
    （3）writeTmpXlsx处理单个sheet
    （4）增加parseRuleXls函数，函数中循环处理*args中的sheetname，分别执行parseSingleSheet和writeTmpXlsx，
    中间表存储在__init__的filepath中
    '''

    #20150430修改，支持调用时传入多个需要处理的sheetname
##    def __init__(self,filepath,*args):
##        self.filepath=os.path.dirname(filepath)          #存储中间表路径
##        self.lstSheets=[]
##        self.workbook=xlrd.open_workbook(filepath)
##        lstSheetnamesInWorkbook=self.workbook.sheet_names()
##        for sheetname in args:
##            if sheetname in lstSheetnamesInWorkbook:
##                self.lstSheets.append(self.workbook.sheet_by_name(sheetname))
##        self.__lstFuncName=[]           #存储功能名称
##        self.__dictSwitchNum={}         #存储功能对应的开关数量
##        self.__dictFunc={}              #存储功能名称和命令开关列表的字典
##
##        #20150506 yh 增加字典存储功能名称对应的需求编号、版本号、描述信息
##        self.__dictReqID={}             #存储功能名称对应的需求编号
##        self.__dictVerID={}             #存储功能名称对应的版本号
##        self.__dictDesc={}              #存储功能名称对应的描述信息

    #20150512 yh workbook中所有sheet页依次处理
    def __init__(self,filepath,middleTableSavePath):
        self.filepath=middleTableSavePath          #存储中间表路径
        self.lstSheets=[]
        #self.workbook=xlrd.open_workbook(filepath)
        self.workbook=xlrd.open_workbook(filepath.decode('utf-8'))
        self.lstSheets=self.workbook.sheets()

        self.__lstFuncName=[]           #存储功能名称
        self.__dictSwitchNum={}         #存储功能对应的开关数量
        self.__dictFunc={}              #存储功能名称和命令开关列表的字典

        #20150506 yh 增加字典存储功能名称对应的需求编号、版本号、描述信息
        self.__dictReqID={}             #存储功能名称对应的需求编号
        self.__dictVerID={}             #存储功能名称对应的版本号
        self.__dictDesc={}              #存储功能名称对应的描述信息

        #根据文件名判断规则表语言，0：英文，1：中文，-1：无法判断
        basename=os.path.basename(filepath)
        if u'英文' in basename:
            self.__lang=0
        elif u'中文' in basename:
            self.__lang=1
        else:
            self.__lang=-1

    #20150430增加，获取字段名称在表头中的索引位置
    def getIndexByFieldname(self, sheet, fieldname):
        i=0
        lstCell=sheet.row(0)
        for cell in lstCell:
            if lstCell[i].value == fieldname:
                return i
            i+=1

    def parseRuleXls(self):
        lstFilterCmdNames=[]
        lstCmdNames=[]
        wb=Workbook()
        #删除默认生成的"Sheet"
        wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
        wbWriter=ExcelWriter(workbook=wb)

        strDstFilepath=self.filepath+'\\middle.xlsx'
        #删除已经存在的中间表middle.xlsx，如果中间表已打开，提示关闭后重试
        if os.path.exists(strDstFilepath):
            try:
                os.remove(strDstFilepath)
            ##如何抛出异常
            except WindowsError:
                print strDstFilepath+'文件已打开，请关闭后重试！'
                return False

        for sheet in self.lstSheets:
            #20150522 yh "命令过滤"sheet页特殊处理
            if sheet.name != "命令过滤":
                idx=self.getIndexByFieldname(sheet,"命令名称")
                lstCmdNames=self.__getCmdNames(sheet,idx,lstCmdNames)
                self.parseSingleSheet(sheet)
                self.writeTmpXlsx(wb,sheet.name)
            else:
                lstFilterCmdNames=self.__getCmdNamesInSheet(sheet,0)
                self.writeFilterCmdXlsx(wb,sheet.name,lstFilterCmdNames)

        #将规则表中使用到的命令名称写到"规则表命令"sheet页中
        self.writeUsingCmdXlsx(wb,"规则表命令",lstCmdNames)
        wbWriter.save(filename=strDstFilepath)

        return lstFilterCmdNames

    #20150507修改，支持新增表头字段
    def parseSingleSheet(self,sheet):
        self.__lstFuncName=[]
        self.__dictSwitchNum.clear()
        self.__dictFunc.clear()
        self.__dictReqID.clear()
        self.__dictVerID.clear()
        self.__dictDesc.clear()

        funcColNum=self.getIndexByFieldname(sheet,u'功能名称')
        demandIDColNum=self.getIndexByFieldname(sheet,u'需求编号')
        verIDColNum=self.getIndexByFieldname(sheet,u'版本号')
        descColNum=self.getIndexByFieldname(sheet,u'描述信息')
        cmdColNum=self.getIndexByFieldname(sheet,u'命令名称')

        lstMerged=[]                #存储合并单元格行信息
        lstMergedColNum=[]          #存储合并单元格的初始行号
        lstSwitch=[]                #存储每一个命令名称对应的开关列表
        dictCmdSwitch={}            #存储命令名称及其对应开关列表的字典

        nrows=sheet.nrows
        '''
        获取合并单元格的行信息
        (row,row_range,col,col_range)
        [row,row_range)包括行合并信息
        [col,col_range)包括列合并信息
        '''
        for crange in sheet.merged_cells:
            rlow, rhigh, clow, chigh = crange
            #仅处理“功能名称”列对应的行号信息
            if clow==funcColNum:
                #记录合并单元格跨越的行号
                lstMerged.append([rlow,rhigh])
                #记录合并单元格初始行号
                lstMergedColNum.append(rlow)
        lstMergedColNum.sort()
        lstMerged.sort(key=operator.itemgetter(0))

        nmerged=len(lstMerged)      #合并单元格个数
        nCurRowNumber=1             #当前行号，默认第1行为标题（索引为0），跳过
        nCurMergedNumber=0          #lstMerged的当前索引
        while nCurRowNumber<nrows:
            nSwitchNum=0
            lstSwitch=[]
            dictCmdSwitch={}

            #如果没有合并单元格则不判断当前行号是否在合并单元格列表中
            if nmerged > 0 and nCurRowNumber in lstMergedColNum:
                strFuncName=sheet.cell_value(lstMerged[nCurMergedNumber][0],funcColNum)
                #20150529 跳过空的合并单元格
                if not strFuncName.strip():
                    nCurRowNumber=lstMerged[nCurMergedNumber][1]
                    nCurMergedNumber=nCurMergedNumber+1
                    nmerged=nmerged-1
                    continue
                self.__lstFuncName.append(strFuncName)
                ############################################
                i=nCurRowNumber
                while i < lstMerged[nCurMergedNumber][1]:   #lstMerged[nCurMergedNumber][1]-1为合并单元格的终止行号
                    #j为第一个开关的列号
                    j=cmdColNum+1
                    while j < len(sheet.row(i)):
                        if sheet.row(i)[j].ctype > 0:  #非空单元格
                            nSwitchNum=nSwitchNum+1
                            lstSwitch.append(sheet.cell_value(i,j))       #命令名称对应的开关集
                        j=j+1
                    dictCmdSwitch[sheet.cell_value(i,cmdColNum)]=lstSwitch        #命令名称及其对应开关的字典
                    lstSwitch=[]
                    i=i+1
                ############################################
                self.__dictFunc[strFuncName]=dictCmdSwitch                                      #功能名称对应的命令名称及其开关集合
                self.__dictSwitchNum[strFuncName]=nSwitchNum                                    #功能名称对应的开关数量
                self.__dictReqID[strFuncName]=sheet.cell_value(nCurRowNumber,demandIDColNum)    #功能名称对应的需求编号
                self.__dictVerID[strFuncName]=sheet.cell_value(nCurRowNumber,verIDColNum)       #功能名称对应的版本号
                self.__dictDesc[strFuncName]=sheet.cell_value(nCurRowNumber,descColNum)         #功能名称对应的描述信息
                nCurRowNumber=lstMerged[nCurMergedNumber][1]
                nCurMergedNumber=nCurMergedNumber+1
                nmerged=nmerged-1
            else:
                strFuncName=sheet.cell_value(nCurRowNumber,funcColNum)
                #20150529 跳过空的合并单元格
                if not strFuncName.strip():
                    nCurRowNumber=nCurRowNumber+1
                    continue
                self.__lstFuncName.append(strFuncName)
                j=cmdColNum+1
                while j < len(sheet.row(nCurRowNumber)):
                    if sheet.row(nCurRowNumber)[j].ctype > 0:
                        nSwitchNum=nSwitchNum+1
                        lstSwitch.append(sheet.cell_value(nCurRowNumber,j))
                    j=j+1
                dictCmdSwitch[sheet.cell_value(nCurRowNumber,cmdColNum)]=lstSwitch
                self.__dictFunc[strFuncName]=dictCmdSwitch
                self.__dictSwitchNum[strFuncName]=nSwitchNum
                self.__dictReqID[strFuncName]=sheet.cell_value(nCurRowNumber,demandIDColNum)
                self.__dictVerID[strFuncName]=sheet.cell_value(nCurRowNumber,verIDColNum)
                self.__dictDesc[strFuncName]=sheet.cell_value(nCurRowNumber,descColNum)
                nCurRowNumber=nCurRowNumber+1

    #20150601 新增处理异常输入代码
    #将规则表中的开关分解为查询字段及查询条件
    def __parseCondition(self,strCmd,strSwitch):
        lstRstStr=[]        #保存中间表的分解字段

        #须保证strSwitch为字符串
        if isinstance(strSwitch,str):
            lstRstStr.append("")
            lstRstStr.append("")
            lstRstStr.append("")
        else:
            regexp=r'([><!]=?|=)'
            prog=re.compile(regexp)
            lstSplitStr=prog.split(strSwitch)

            #不能解析为3部分
            if len(lstSplitStr) != 3:
                if u'非空'==strSwitch:
                    lstRstStr.append("")
                    lstRstStr.append("")
                    lstRstStr.append("NotNull")
                elif u'非空' in strSwitch:        #包含“非空”的字符串
                    prog=re.compile(u'(.*?)(非空)',re.UNICODE)
                    lstRstStr.append(prog.search(strSwitch).group(1))
                    lstRstStr.append("")
                    lstRstStr.append("NotNull")
                else:
                    #既不能分解为3部分，又不包含"非空"字符串，则填充空字符串
                    lstRstStr.append("")
                    lstRstStr.append("")
                    lstRstStr.append("")
            else:
                if "SHOW SOFTWARE PARAMETER"==strCmd:
                    #20150528 yh 根据语言类型确定查询字段
                    if self.__lang==0:
                        lstRstStr.append("CurrentValue")
                        lstRstStr.append("ParameterID="+lstSplitStr[0])
                    #__lang非0表示为中文，没有判断__lang=-1的情况
                    else:
                        lstRstStr.append("参数当前值")
                        lstRstStr.append("软件参数ID="+lstSplitStr[0])
                    #lstRstStr.append("ParameterID="+lstSplitStr[0])
                    lstRstStr.append(lstSplitStr[2])
                else:
                    lstRstStr.append(lstSplitStr[0])
                    lstRstStr.append("")
                    lstRstStr.append(lstSplitStr[2])
                    #lstRstStr.append(lstSplitStr[2].capitalize())

                #20150529 处理类似"262226 !=0"开关，将预期结果保存为!0
                if lstSplitStr[1].strip() == '!=':
                    lstRstStr[2]='!'+lstSplitStr[2].strip()

        return lstRstStr

##    #将规则表中的开关分解为查询字段及查询条件
##    def __parseCondition(self,strCmd,strSwitch):
##        regexp=r'([><!]=?|=)'
##        prog=re.compile(regexp)
##        ##须保证strSwitch为字符串，否则不进行解析
##        lstSplitStr=prog.split(strSwitch)
##        lstRstStr=[]        #保存中间表的分解字段
##
##        #不能解析为3部分，则认为解析字符串出错
####        if len(lstSplitStr) != 3:
####            return None
##
##        if "SHOW SOFTWARE PARAMETER"==strCmd:
##            #20150528 yh 根据语言类型确定查询字段
##            if self.__lang==0:
##                lstRstStr.append("CurrentValue")
##                lstRstStr.append("ParameterID="+lstSplitStr[0])
##            #__lang非0表示为中文，没有判断__lang=-1的情况
##            else:
##                lstRstStr.append("参数当前值")
##                lstRstStr.append("软件参数ID="+lstSplitStr[0])
##            #lstRstStr.append("ParameterID="+lstSplitStr[0])
##            lstRstStr.append(lstSplitStr[2])
##        elif u'非空'==strSwitch:
##            lstRstStr.append("")
##            lstRstStr.append("")
##            lstRstStr.append("NotNull")
##        elif u'非空' in strSwitch:        #包含“非空”的字符串
##            prog=re.compile(u'(.*?)(非空)',re.UNICODE)
##            lstRstStr.append(prog.search(strSwitch).group(1))
##            lstRstStr.append("")
##            lstRstStr.append("NotNull")
##        else:
##            lstRstStr.append(lstSplitStr[0])
##            lstRstStr.append("")
##            lstRstStr.append(lstSplitStr[2])
##            #lstRstStr.append(lstSplitStr[2].capitalize())
##
##        #20150529 处理类似"262226 !=0"开关，将预期结果保存为!0
##        if len(lstSplitStr)>1 and lstSplitStr[1].strip() == '!=':
##            lstRstStr[2]='!'+lstSplitStr[2]
##
##        return lstRstStr

    #获取sheet页中"命令名称"列的值
    def __getCmdNames(self, sheet, idx, lstCmdNames):
        #lstCmdNames=[]
        i=1
        while i < sheet.nrows:
            if sheet.cell_type(i,idx)>0:
                lstCmdNames.append(sheet.cell_value(i,idx))
            i=i+1

        #20150422 yh 去除重复元素
        lstCmdNames=list(set(lstCmdNames))

        return lstCmdNames

    #20150522 yh 提取sheet页中的index列的命令
    def __getCmdNamesInSheet(self, sheet, index):
        lstCmdNames=[]
        i=0
        while i < sheet.nrows:
            if sheet.cell_type(i,index)>0:
                lstCmdNames.append(sheet.cell_value(i,index))
            i=i+1

        #去除重复元素
        lstCmdNames=list(set(lstCmdNames))

        return lstCmdNames

    #将规则表转换为程序处理的中间表，xlsx格式
    #def writeTmpXlsx(self,strDstFilepath):
    def writeTmpXlsx(self,wb,strSheetName):
        #如果中间表已经存在，需删除
        ##如果中间表已打开，须提示先关闭
##        if os.path.exists(strDstFilepath):
##            try:
##                os.remove(strDstFilepath)
##            ##如何处理异常
##            except OSError,e:
##                print e

##        wb=Workbook()
##        wbWriter=ExcelWriter(workbook=wb)

        ws=wb.create_sheet()
        ws.title=strSheetName
        #ws=wb.worksheets[0]     #取得wb的第一个工作表
        #ws.title="test"

        ncol=1
        nrow=1
        for funcName in self.__lstFuncName:
            #ncol=1
            ncol=0
            ws.cell(row=nrow,column=ncol).value=self.__dictReqID[funcName]                  #写入需求编号
            ncol+=1
            ws.cell(row=nrow,column=ncol).value=self.__dictVerID[funcName]                  #写入版本号
            ncol+=1
            ws.cell(row=nrow,column=ncol).value=funcName                                    #写入功能名称
            ncol+=1
            ws.cell(row=nrow,column=ncol).value=self.__dictDesc[funcName]                   #写入描述信息
            ncol+=1
            #20150601 插入flag列，表达开关之间的关系，0：与，1：非，2：或，默认为0
            nFlagCol=ncol
            #20150601 'RFSP功能'特殊处理
            if funcName==u'RFSP功能':
                ws.cell(row=nrow,column=ncol).value=1
            else:
                ws.cell(row=nrow,column=ncol).value=0
            ncol+=1
            #20150529 保存开关数量对应的列号
            nSwitchNumCol=ncol
            ws.cell(row=nrow,column=ncol).value=self.__dictSwitchNum[funcName]              #写入开关数量
            ncol+=1
            d=self.__dictFunc[funcName]
            for k in d.keys():                                                              #k是命令名称
                for strSwitch in d[k]:
                    #20150529 处理类似"Barring of All PS Services=Yes | Forbids to Visit HPLMN APN=Yes | Forbids to Visit VPLMN APN=Yes"的开关
                    lstSwitch=strSwitch.split('|')
                    if len(lstSwitch)>1:
                        #20150529 更新falg和开关数量
                        ws.cell(row=nrow,column=nFlagCol).value=2
                        ws.cell(row=nrow,column=nSwitchNumCol).value=self.__dictSwitchNum[funcName]+len(lstSwitch)-1
                    for subSwitch in lstSwitch:
                        ws.cell(row=nrow,column=ncol).value=re.sub(r'\s+',' ',k).replace(' ','_')       #写入命令名称，将空格用下划线代替
                        #ws.cell(row=nrow,column=ncol).value=k.replace(' ','_')
                        ncol=ncol+1
                        lstRstStr=self.__parseCondition(k,subSwitch)
                        #20150601 如果开关中包括'!'，则更新flag列
##                        if '!' in lstRstStr[2]:
##                            ws.cell(row=nrow,column=nFlagCol).value=1
                        for i in range(3):
                            #20150512 yh 仅对查询字段去除空格
                            #strS=lstRstStr[i].replace(' ','')       #去除空格

                            #20150512 yh 去除字符串首尾的空格
                            strS=lstRstStr[i].strip()

                            #20150416 yh 对查询字段去除字符串间的空格和特殊字符[/\-]
                            if i==0:
                                strS=lstRstStr[i].replace(' ','')
                                strS=re.sub('[/\\\\-]','',strS)
                            ws.cell(row=nrow,column=ncol).value=strS
                            ncol=ncol+1
            nrow=nrow+1

    #20150522 yh 写规则表中的“命令过滤”sheet页
    def writeFilterCmdXlsx(self,wb,strSheetName, lstFilterCmdNames):
        ws=wb.create_sheet()
        ws.title=strSheetName

        ncol=1
        nrow=1
        i=0
        for funcName in lstFilterCmdNames:
            ws.cell(row=nrow,column=ncol).value=lstFilterCmdNames[i]
            nrow+=1
            i+=1

    #20150522 yh 把规则表中的所有命令名称写入到"规则表命令"sheet页
    def writeUsingCmdXlsx(self,wb,strSheetName, lstCmdNames):
        ws=wb.create_sheet()
        ws.title=strSheetName

        ncol=1
        nrow=1
        i=0
        while(i<len(lstCmdNames)):
            ws.cell(row=nrow,column=ncol).value=lstCmdNames[i]
            nrow+=1
            i+=1

##    #将规则表转换为程序处理的中间表，xls格式
##    def writeTmpXls(self,strDstFilepath):
##        #如果中间表已经存在，需删除
##        if os.path.exists(strDstFilepath):
##            try:
##                os.remove(strDstFilepath)
##            ##如何处理异常
##            except OSError,e:
##                print e
##
##        wb = xlwt.Workbook()
##        ws = wb.add_sheet('Test Sheet')
##        ncol=0
##        nrow=0
##        for funcName in self.__lstFuncName:
##            ncol=0
##            ws.write(nrow,ncol,funcName)                            #写入功能名称
##            ncol=ncol+1
##            ws.write(nrow,ncol,self.__dictSwitchNum[funcName])      #写入开关数量
##            ncol=ncol+1
##            d=self.__dictFunc[funcName]
##            for k in d.keys():                                      #k是命令名称
##                for strSwitch in d[k]:
##                    #20150529 处理类似"Barring of All PS Services=Yes | Forbids to Visit HPLMN APN=Yes | Forbids to Visit VPLMN APN=Yes"的开关
##                    lstSwitch=strSwitch.split('|')
##                    for subSwitch in lstSwitch:
##                        ws.write(nrow,ncol,k.replace(' ','_'))           #写入命令名称，将空格用下划线代替
##                        ncol=ncol+1
##                        lstRstStr=self.__parseCondition(k,subSwitch.strip())
##                        ##如果lstRstStr为空列表，需特殊处理
##                        for i in range(3):
##                            ws.write(nrow,ncol,lstRstStr[i])
##                            ncol=ncol+1
##            nrow=nrow+1
##
##        #把规则表中的所有命令名称写入到"cmdNames"sheet
##        ws = wb.add_sheet('cmdNames')
##        lstCmdNames=self.__getCmdNames()
##        for i in range(len(lstCmdNames)):
##            ws.write(i,0,lstCmdNames[i])
##
##        wb.save(strDstFilepath)

if __name__ == '__main__':
    #filepath=r'D:\customtool\data\H3G.xlsx'
    #filepath=r'D:\customtool\data\英文规则表.xlsx'
    #filepath=r'D:\customtool\data\test4功能开关-英文规则（EMM）.xlsx'
    #filepath=r'D:\customtool\data\功能开关-英文规则（EMM）.xlsx'
    filepath=r'D:\customtool\data\合并中文规则表-去除不支持部分.xlsx'
    #filepath=r'D:\customtool\se.xlsx'
    #sheetname='Sheet3'

    #xls=xlsOP(filepath,"SGSN","MME","公共")
    xls=xlsOP(filepath)
    xls.parseRuleXls()
#    xls.writeTmpXls(r'D:\customtool\example.xls')
    #xls.writeTmpXlsx(r'D:\customtool\example.xlsx')
    #xls.getCmdNames()

    #print "done"

    pass